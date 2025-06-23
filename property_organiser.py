import pandas as pd
from openpyxl import load_workbook
import os

from constants import Headings
class PropertyOrganiser:
    def __init__(self, current_search):
        self.current_search = current_search
        self.df = None
        if os.path.exists(current_search):
            self.df = pd.read_excel(self.current_search)
        else:
            self.df = pd.DataFrame(columns=Headings.DATAFRAME.value)
            self.df.to_excel(current_search, index=False)

    def add_from_api(self, data_frame):
        combined_df = pd.concat([self.df, data_frame], ignore_index=True)
        deduplicated_df = combined_df.drop_duplicates(subset=['URL'])

        # Update and save
        self.df = deduplicated_df
        self.df.to_excel(self.current_search, index=False)

    def add_from_scraper(self, data_frame):
        self.df = pd.merge(self.df, data_frame, on="URL", how="left")
        self.df.to_excel(self.current_search, index=False)

    def delete_from_scraper(self, rows_to_delete):
        self.df = self.df[~self.df['URL'].isin(rows_to_delete)]
        self.df.to_excel(self.current_search, index=False)

    def get_urls(self, link_search):
        wb = load_workbook(link_search)
        self.ws = wb.active
        urls = []
        for row in self.ws.iter_rows(min_row=1, max_col=1):
            cell = row[0]
            if cell.hyperlink:
                urls.append(cell.hyperlink.target)
            else:
                urls.append(cell.value)
        return urls

    def convert_to_csv(self, csv):
        self.df.to_csv(csv, index=False)

    def review_df(self):
        rows_to_keep = []
        rows_to_delete = []

        for idx, row in df.iterrows():
            print(f"\nRow {idx}:")
            print(row["URL"])

            action = input("Enter 'd' to delete, 'n' to add a note and keep, or just Enter to keep as-is: ").strip().lower()

            if action == 'd':
                print("Row marked for deletion.")
                rows_to_delete.append(idx)
            elif action == 'n':
                note = input("Enter your note: ").strip()
                df.at[idx, 'Note'] = note
                rows_to_keep.append(idx)
            else:
                rows_to_keep.append(idx)

        # Drop rows marked for deletion
        df = df.drop(rows_to_delete).reset_index(drop=True)